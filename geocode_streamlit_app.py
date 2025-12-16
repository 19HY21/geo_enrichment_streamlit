# -*- coding: utf-8 -*-
"""
Created on Thu Jun  6 14:00:00 2024
Geo Enrichment Tool (Streamlit)
- geo_logic/core を利用し、ブラウザから郵便番号/住所突合とジオコーディングを実行
- Streamlit Cloud でも動作するよう軽量構成
"""

import io
import base64
import os
import sys
from datetime import datetime
from typing import List, Tuple

import pandas as pd
import streamlit as st

BASE_DIR = os.path.dirname(__file__)
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from geo_logic import core as logic  # noqa: E402

# マスタパスをリポジトリ内 data に固定したい場合はコメントアウトを外してください
# logic.MASTER_PATH = os.path.join(BASE_DIR, "data", "zipcode_localgoverment_mst.xlsx")

CACHE_DIR = logic.CACHE_DIR
OUTPUT_SUFFIX = logic.OUTPUT_SUFFIX
BATCH_SIZE_DEFAULT = 1_000  # ジオコーディング一括処理サイズ（住所数）
attach_master_by_address = logic.attach_master_by_address  # 住所突合
attach_master_by_zip = logic.attach_master_by_zip  # 郵便番号突合
geocode_addresses = logic.geocode_addresses  # ジオコーディング
load_cache = logic.load_cache  # キャッシュ読込
read_master = logic.read_master  # マスタ読込
save_cache = logic.save_cache  # キャッシュ保存
add_geocode_columns = logic.add_geocode_columns  # 緯度経度列付与
normalize_address = logic.normalize_address  # 住所正規化


def _log(log_box, msg: str):
    # ログ追記
    ts = datetime.now().strftime("%H:%M:%S")
    logs = st.session_state.setdefault("logs", [])
    logs.append(f"[{ts}] {msg}")
    log_box.write("\n".join(logs))


def _run_pipeline(
    # 主要処理パイプライン
    df_input: pd.DataFrame,
    sheet_name: str,
    file_kind: str,
    zip_cols: List[str],
    addr_cols: List[str],
    xls_for_copy: pd.ExcelFile,
    log_box,
    base_name: str,
    batch_size: int,
    geocode_enabled: bool = True,
    uploaded_cache: dict | None = None,
    process_mask: pd.Series | None = None,
    chunk_offset: int = 0,
):
    # ダウンロードUIを進捗直下に配置
    st.session_state.setdefault("addr_chunk_downloads", [])
    st.session_state.setdefault("geo_chunk_downloads", [])
    st.session_state.setdefault("logs", [])
    progress = st.progress(0)
    status = st.empty()
    live_download_section = st.container()
    with live_download_section:
        st.markdown("住所/ジオのチャンクダウンロード")

    weights = {"zip": 20, "addr": 20, "geo": 55, "out": 5}
    enabled_phases = ["zip"]
    if addr_cols:
        enabled_phases.append("addr")
    if addr_cols and geocode_enabled:
        enabled_phases.append("geo")
    enabled_phases.append("out")
    total_weight = sum(weights[p] for p in enabled_phases)

    def set_progress(phase: str, pct: float, text: str):
        # 進捗バー更新
        offset = 0
        for p in ["zip", "addr", "geo", "out"]:
            if p == phase:
                break
            if p in enabled_phases:
                offset += weights[p]
        if phase not in enabled_phases:
            return
        w = weights[phase]
        overall = (offset + w * (pct / 100.0)) / total_weight * 100.0
        progress.progress(min(max(int(overall), 0), 100))
        status.write(text)

    def prog_bar(phase, pct, text):
        set_progress(phase, pct, text)

    _log(log_box, "マスタ読込開始")
    master_df = read_master()
    _log(log_box, "マスタ読込完了")
    total_rows_all = len(df_input)
    # process_mask は住所突合をスキップする行のマークとしてのみ利用し、全体処理は全行を対象とする
    df_proc_in = df_input.copy()
    _log(
        log_box,
        f"入力件数: {total_rows_all} / 住所突合対象件数: "
        f"{len(df_input) if process_mask is None else process_mask.sum()}",
    )

    cols_needed = list(dict.fromkeys(zip_cols + addr_cols))
    df_work = df_proc_in[cols_needed].copy() if cols_needed else df_proc_in.copy()
    used_zip_codes = set()
    used_master_idx = set()

    if df_work.empty:
        _log(log_box, "処理対象の行がありません（全件Parquet由来など）。")
        out_base = base_name or "output"
        df_out_merge = df_input.copy()
        for helper_col in ["__merge_key", "__is_parquet"]:
            if helper_col in df_out_merge.columns:
                df_out_merge = df_out_merge.drop(columns=[helper_col])
        df_input_clean = df_input.drop(columns=["__merge_key", "__is_parquet"], errors="ignore")
        if file_kind == "excel" and xls_for_copy is not None:
            buf = _build_excel_output(
                xls_for_copy, sheet_name, df_input_clean, df_out_merge, read_master(), set(), set()
            )
            fname = f"{out_base}{OUTPUT_SUFFIX}.xlsx"
        else:
            buf = _build_csv_output(df_out_merge)
            fname = f"{out_base}{OUTPUT_SUFFIX}.csv"
        return buf, fname, df_out_merge, os.path.join(CACHE_DIR, "streamlit_local_cache.parquet")

    # 郵便番号突合
    if zip_cols:
        _log(log_box, "郵便番号突合開始")

        def zip_prog(done, total, detail):
            pct = done / max(total, 1) * 100
            prog_bar("zip", pct, f"[zip] {detail}")

        df_work = attach_master_by_zip(
            df_work, master_df, zip_cols, progress=zip_prog, used_zip_codes=used_zip_codes
        )
        prog_bar("zip", 100, "[zip] 完了")
        _log(log_box, f"郵便番号突合完了: {len(used_zip_codes)}件")

    # 住所突合（チャンク＋オンディスク保存）: 突合済みParquetと一致した行はスキップ、未一致のみ処理
    if addr_cols:
        addr_mask = process_mask
        df_addr_target = df_work if addr_mask is None else df_work.loc[addr_mask].copy()
        total_rows = len(df_addr_target)
        _log(log_box, f"住所突合開始 / 対象件数: {total_rows}件")
        if total_rows > 0:
            chunk_size = 1000
            addr_chunks = []
            processed = 0
            chunk_dir = os.path.join(CACHE_DIR, "addr_chunks")
            os.makedirs(chunk_dir, exist_ok=True)

            for start in range(0, total_rows, chunk_size):
                end = min(start + chunk_size, total_rows)
                chunk = df_addr_target.iloc[start:end].copy()
                chunk = attach_master_by_address(
                    chunk, master_df, addr_cols, progress=None, used_master_idx=used_master_idx
                )
                addr_chunks.append(chunk)
                chunk_fname = f"{base_name or 'output'}_addr_chunk_{start+1+chunk_offset}_{end+chunk_offset}.parquet"
                chunk_path = os.path.join(chunk_dir, chunk_fname)
                chunk.to_parquet(chunk_path, index=False)
                try:
                    buf = io.BytesIO()
                    chunk.to_parquet(buf, index=False)
                    buf.seek(0)
                    b64 = base64.b64encode(buf.getvalue()).decode()
                    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{chunk_fname}">住所チャンク {start+1+chunk_offset}-{end+chunk_offset} をダウンロード (Parquet)</a>'
                    st.session_state["addr_chunk_downloads"].append(href)
                    st.markdown(href, unsafe_allow_html=True)
                except Exception:
                    pass
                processed = end
                pct = processed / max(total_rows, 1) * 100
                prog_bar("addr", pct, f"[addr] {processed}/{total_rows} ({pct:.1f}%)")
                _log(log_box, f"[addr] 進捗 {processed}/{total_rows} ({pct:.1f}%)")

            df_addr_out = pd.concat(addr_chunks).sort_index()
            # 処理した行だけ反映
            df_work.loc[df_addr_out.index, df_addr_out.columns] = df_addr_out
            prog_bar("addr", 100, "[addr] 完了")
            _log(log_box, f"住所突合完了 使用行: {len(used_master_idx)}件")
        else:
            _log(log_box, "住所突合スキップ（対象行なし）")

    # ジオコーディング
    os.makedirs(CACHE_DIR, exist_ok=True)
    local_cache_path = os.path.join(CACHE_DIR, "streamlit_local_cache.parquet")
    cache = load_cache(local_cache_path)
    if uploaded_cache:
        cache.update(uploaded_cache)
    _log(log_box, f"キャッシュ読込: ローカル{len(cache)}件")

    geo_results = {}
    if addr_cols and geocode_enabled:
        _log(log_box, "ジオコーディング開始")
        prog_bar("geo", 0, "[geo] 処理開始")
        all_addrs = []
        for col in addr_cols:
            all_addrs.extend(df_work[col].dropna().tolist())
        unique_addrs = [a for a in pd.Series(all_addrs).dropna().unique().tolist() if normalize_address(a)]
        total_unique = len(unique_addrs)
        _log(log_box, f"ユニーク住所数: {total_unique}件 / バッチサイズ: {batch_size}")

        overall_done = 0
        for start in range(0, total_unique, batch_size):
            end = min(start + batch_size, total_unique)
            chunk = unique_addrs[start:end]

            def geo_prog(done, total, kind):
                now_done = overall_done + done
                pct = now_done / max(total_unique, 1) * 100
                prog_bar("geo", pct, f"[geo] {now_done}/{total_unique} ({pct:.1f}%)")

            def geo_cache_save(c):
                save_cache(local_cache_path, c)

            _log(log_box, f"バッチ処理: {start+1}〜{end}件目")
            chunk_results, cache_hit, new_count = geocode_addresses(
                chunk,
                user_agent="GeoGUI_streamlit",
                cache=cache,
                progress_cb=geo_prog,
                cache_save_cb=geo_cache_save,
            )
            geo_results.update(chunk_results)
            save_cache(local_cache_path, cache)
            try:
                geo_df = pd.DataFrame(
                    [
                        {"address": k, "lat": v[0], "lon": v[1], "flag": v[2]}
                        for k, v in chunk_results.items()
                    ]
                )
                geo_chunk_fname = f"{base_name or 'output'}_geo_chunk_{start+1+chunk_offset}_{end+chunk_offset}.parquet"
                geo_bytes = io.BytesIO()
                geo_df.to_parquet(geo_bytes, index=False)
                geo_bytes.seek(0)
                b64 = base64.b64encode(geo_bytes.getvalue()).decode()
                href = f'<a href="data:application/octet-stream;base64,{b64}" download="{geo_chunk_fname}">ジオコードチャンク {start+1+chunk_offset}-{end+chunk_offset} をダウンロード (Parquet)</a>'
                st.session_state["geo_chunk_downloads"].append(href)
                st.markdown(href, unsafe_allow_html=True)
            except Exception:
                pass
            overall_done = end
            _log(log_box, f"バッチ完了 cache_hit={cache_hit} 新規={new_count} 累計={overall_done}/{total_unique}")

    df_work = add_geocode_columns(df_work, addr_cols, geo_results)
    if not (addr_cols and geocode_enabled):
        _log(log_box, "ジオコーディングはスキップ（住所未選択または緯度経度付与オフ）")

    if geocode_enabled and addr_cols:
        save_cache(local_cache_path, cache)

    out_base = base_name or "output"
    df_out_merge = df_input.copy()
    for col in df_work.columns:
        df_out_merge.loc[df_work.index, col] = df_work[col]
    for helper_col in ["__merge_key", "__is_parquet"]:
        if helper_col in df_out_merge.columns:
            df_out_merge = df_out_merge.drop(columns=[helper_col])

    df_input_clean = df_input.drop(columns=["__merge_key", "__is_parquet"], errors="ignore")

    if file_kind == "excel":
        buf = _build_excel_output(
            xls_for_copy, sheet_name, df_input_clean, df_out_merge, master_df, used_zip_codes, used_master_idx
        )
        fname = f"{out_base}{OUTPUT_SUFFIX}.xlsx"
    else:
        buf = _build_csv_output(df_out_merge)
        fname = f"{out_base}{OUTPUT_SUFFIX}.csv"

    prog_bar("out", 50, "[out] 生成中")
    _log(log_box, f"出力生成完了: {fname}")
    prog_bar("out", 100, "完了")

'''
    return buf, fname, df_out_merge, local_cache_path
'''


def _build_excel_output(
    xls: pd.ExcelFile,
    sheet_name: str,
    original_df: pd.DataFrame,
    processed_df: pd.DataFrame,
    master_df: pd.DataFrame,
    used_zip_codes: set,
    used_master_idx: set,
):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name in xls.sheet_names:
            df_sheet = xls.parse(name, dtype=str)
            df_sheet.to_excel(writer, sheet_name=name, index=False)
        original_df.to_excel(writer, sheet_name=sheet_name, index=False)
        processed_df.to_excel(writer, sheet_name=f"{sheet_name}{OUTPUT_SUFFIX}", index=False)
        master_df.to_excel(writer, sheet_name="master", index=False)
        try:
            from openpyxl.styles import PatternFill

            ws = writer.book["master"]
            highlight_idx = set(used_master_idx) if used_master_idx else set()
            if used_zip_codes:
                matched = master_df[master_df["郵便番号"].isin(used_zip_codes)]
                highlight_idx.update(matched.index.tolist())
            if highlight_idx:
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for idx in highlight_idx:
                    row_num = idx + 2
                    for col_num in range(1, len(master_df.columns) + 1):
                        ws.cell(row=row_num, column=col_num).fill = fill
        except Exception:
            pass
    buf.seek(0)
    return buf


def _build_csv_output(processed_df: pd.DataFrame):
    buf = io.BytesIO()
    processed_df.to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)
    return buf


def _load_input(uploaded_file) -> Tuple[str, pd.DataFrame, pd.ExcelFile, str, str]:
    file_bytes = uploaded_file.read()
    file_kind = "excel" if uploaded_file.name.lower().endswith(("xlsx", "xls")) else "csv"
    if file_kind == "excel":
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
        sheet = xls.sheet_names[0]
        df = xls.parse(sheet, dtype=str)
        sheet_name = sheet
    else:
        xls = None
        sheet_name = "data"
        df = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
    base_name = os.path.splitext(uploaded_file.name)[0]
    return file_kind, df, xls, sheet_name, base_name


def main():
    st.set_page_config(page_title="Geo Enrichment Tool", layout="wide")
    st.title("Geo Enrichment Tool")
    st.caption(
        "住所・郵便番号から、日本郵政マスタを用いて地域情報と緯度経度を付与し、データを生成するアプリです。"
    )

    uploaded = st.file_uploader("入力ファイルを選択 (Excel/CSV)", type=["csv", "xlsx", "xls"])
    parquet_uploader = st.file_uploader(
        "突合済みParquetをアップロード（任意・複数可）",
        type=["parquet"],
        key="parquet_uploader",
        accept_multiple_files=True,
    )

    st.session_state.setdefault("addr_chunk_downloads", [])
    st.session_state.setdefault("geo_chunk_downloads", [])
    st.session_state.setdefault("result_file", None)
    st.session_state.setdefault("cache_file", None)

    if uploaded or parquet_uploader:
        df_parquet = None
        parquet_base_name = None
        parquet_files = parquet_uploader if parquet_uploader else []
        if parquet_files:
            pq_frames = []
            for f in parquet_files:
                pq_frames.append(pd.read_parquet(io.BytesIO(f.read())))
            df_parquet = pd.concat(pq_frames, ignore_index=True) if pq_frames else None
            parquet_base_name = os.path.splitext(parquet_files[0].name)[0]
        process_mask = None
        chunk_offset = 0
        if uploaded:
            file_kind, df_input, xls_for_copy, sheet_name, base_name = _load_input(uploaded)
            if file_kind == "excel" and xls_for_copy is not None:
                sheet_name = st.selectbox("シート名を選択", options=xls_for_copy.sheet_names, index=0)
                df_input = xls_for_copy.parse(sheet_name, dtype=str)
        elif df_parquet is not None:
            df_input = df_parquet
            file_kind = "parquet"
            xls_for_copy = None
            sheet_name = "data"
            base_name = parquet_base_name or "parquet_input"

        zip_cols = st.multiselect("郵便番号列を選択", options=df_input.columns.tolist())
        addr_cols = st.multiselect("住所列を選択", options=df_input.columns.tolist())
        geocode_enabled = st.checkbox("緯度経度を付与する", value=False)
        cache_uploader = st.file_uploader(
            "キャッシュParquetをアップロード（任意・複数可）", type=["parquet"], accept_multiple_files=True
        )

        uploaded_cache = None
        if cache_uploader:
            try:
                frames = []
                for f in cache_uploader:
                    frames.append(pd.read_parquet(io.BytesIO(f.read())))
                cache_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
                required_cols = {"address", "lat", "lon", "flag"}
                if not required_cols.issubset(set(cache_df.columns)):
                    raise ValueError("必要なカラム(address, lat, lon, flag)が見つかりません")
                uploaded_cache = {
                    str(row["address"]): (row["lat"], row["lon"], row["flag"])
                    for _, row in cache_df.iterrows()
                    if pd.notna(row.get("address"))
                }
            except Exception as e:
                st.warning(f"キャッシュParquetの読み込みに失敗しました: {e}")

        # Parquetがある場合は住所列キーでParquetを優先マージ（未一致のみ処理）
        if df_parquet is not None and addr_cols:
            missing_base = [c for c in addr_cols if c not in df_input.columns]
            missing_parquet = [c for c in addr_cols if c not in df_parquet.columns]
            if missing_base or missing_parquet:
                st.warning(
                    f"Parquet優先マージをスキップしました。欠損列: "
                    f"ベース={missing_base or 'なし'}, Parquet={missing_parquet or 'なし'}"
                )
            else:
                def _build_key(df):
                    def _to_key(val):
                        if pd.isna(val):
                            return ""
                        if isinstance(val, float) and val.is_integer():
                            val = int(val)
                        return normalize_address(str(val))

                    parts = [df[col].apply(_to_key) for col in addr_cols]
                    key = parts[0]
                    for p in parts[1:]:
                        key = key + "||" + p
                    return key

                base_df = df_input.copy()
                pq_df = df_parquet.copy()
                base_df["_merge_key"] = _build_key(base_df)
                pq_df["_merge_key"] = _build_key(pq_df)

                pq_df = pq_df[pq_df["_merge_key"] != ""]
                pq_df = pq_df.drop_duplicates("_merge_key", keep="first")

                for col in pq_df.columns:
                    if col not in base_df.columns:
                        base_df[col] = None

                base_df = base_df.set_index("_merge_key")
                pq_df = pq_df.set_index("_merge_key")

                base_keys = set(base_df.index[base_df.index != ""])
                pq_keys = set(pq_df.index[pq_df.index != ""])
                common_keys = base_keys & pq_keys
                base_only = base_keys - pq_keys
                pq_only = pq_keys - base_keys

                aligned = pq_df.reindex(base_df.index)
                base_df.update(aligned)

                base_df["__is_parquet"] = base_df.index.isin(common_keys)
                df_input = base_df.reset_index().rename(columns={"index": "__merge_key"})

                st.info(
                    f"キー件数 ベース={len(base_keys)} / Parquet={len(pq_keys)} / 共通={len(common_keys)} / "
                    f"ベースのみ={len(base_only)} / Parquetのみ={len(pq_only)}"
                )

                process_mask = ~df_input["__is_parquet"]
                chunk_offset = int(df_input["__is_parquet"].sum())

        run_clicked = st.button("実行 / 再実行", type="primary")

        if run_clicked:
            log_box = st.empty()
            buf, fname, df_out, cache_path = _run_pipeline(
                df_input=df_input,
                sheet_name=sheet_name,
                file_kind=file_kind,
                zip_cols=zip_cols,
                addr_cols=addr_cols,
                xls_for_copy=xls_for_copy,
                log_box=log_box,
                base_name=base_name,
                batch_size=BATCH_SIZE_DEFAULT,
                geocode_enabled=geocode_enabled,
                uploaded_cache=uploaded_cache,
                process_mask=process_mask,
                chunk_offset=chunk_offset,
            )

            st.session_state["result_file"] = {
                "data": buf.getvalue(),
                "name": fname,
            }
            if os.path.exists(cache_path):
                with open(cache_path, "rb") as f:
                    st.session_state["cache_file"] = {
                        "data": f.read(),
                        "name": os.path.basename(cache_path),
                    }

    # 進捗エリア下に常時ダウンロード表示（チャンクはリンクのみ）
    download_section = st.container()
    with download_section:
        if st.session_state.get("addr_chunk_downloads"):
            st.markdown("住所突合チャンクのダウンロード（処理中も利用可）")
            for item in st.session_state["addr_chunk_downloads"]:
                st.markdown(item, unsafe_allow_html=True)
        if st.session_state.get("geo_chunk_downloads"):
            st.markdown("ジオコーディングチャンクのダウンロード（処理中も利用可）")
            for item in st.session_state["geo_chunk_downloads"]:
                st.markdown(item, unsafe_allow_html=True)
        if st.session_state.get("result_file"):
            st.download_button(
                label="結果データをダウンロード",
                data=st.session_state["result_file"]["data"],
                file_name=st.session_state["result_file"]["name"],
                mime="application/octet-stream",
                key="result_download_persistent",
            )
        if st.session_state.get("cache_file"):
            st.download_button(
                label="キャッシュParquetをダウンロード（次回再利用用）",
                data=st.session_state["cache_file"]["data"],
                file_name=st.session_state["cache_file"]["name"],
                mime="application/octet-stream",
                key="cache_download_persistent",
            )

if __name__ == "__main__":
    main()
