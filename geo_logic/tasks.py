# -*- coding: utf-8 -*-
"""
Batch geocoding task helpers (Streamlit/FastAPI 共通ロジック)

- 入力ファイルを読み込み（CSV/Excel対応）
- 郵便番号/住所突合
- ユニーク化した住所をバッチサイズごとにジオコーディング
- キャッシュはジョブ単位で保存し、途中で落ちても再開しやすい形にする
- 出力は元ファイル拡張子に合わせて CSV / Excel を生成
"""
from __future__ import annotations

import io
import json
import os
from typing import Dict, List, Optional, Tuple

import pandas as pd

from . import core as logic

CACHE_DIR = logic.CACHE_DIR
OUTPUT_SUFFIX = logic.OUTPUT_SUFFIX
BATCH_SIZE_DEFAULT = logic.BATCH_SIZE_DEFAULT


def _build_excel_output(
    xls: pd.ExcelFile,
    sheet_name: str,
    original_df: pd.DataFrame,
    processed_df: pd.DataFrame,
    master_df: pd.DataFrame,
    used_zip_codes: set,
    used_master_idx: set,
) -> io.BytesIO:
    """Excel出力をBytesIOに作成。"""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name in xls.sheet_names:
            df_sheet = xls.parse(name, dtype=str)
            df_sheet.to_excel(writer, sheet_name=name, index=False)
        original_df.to_excel(writer, sheet_name=sheet_name, index=False)
        processed_df.to_excel(writer, sheet_name=f"{sheet_name}{OUTPUT_SUFFIX}", index=False)
        master_df.to_excel(writer, sheet_name="master", index=False)
        try:
            from openpyxl.styles import PatternFill

            ws = writer.book["master"]
            highlight_idx = set(used_master_idx) if used_master_idx else set()
            if used_zip_codes:
                matched = master_df[master_df["郵便番号"].isin(used_zip_codes)]
                highlight_idx.update(matched.index.tolist())
            if highlight_idx:
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for idx in highlight_idx:
                    row_num = idx + 2
                    for col_num in range(1, len(master_df.columns) + 1):
                        ws.cell(row=row_num, column=col_num).fill = fill
        except Exception:
            pass
    buf.seek(0)
    return buf


def _build_csv_output(processed_df: pd.DataFrame) -> io.BytesIO:
    buf = io.BytesIO()
    processed_df.to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)
    return buf


def _load_input_file(file_path: str, sheet_name: Optional[str] = None) -> Tuple[str, pd.DataFrame, Optional[pd.ExcelFile], str, str]:
    """
    ファイルパスからデータを読み込み、(name, df, xls, kind, sheet_name) を返す。
    kind: "csv" or "excel"
    """
    name = os.path.basename(file_path)
    lower = name.lower()
    if lower.endswith(".csv"):
        df = pd.read_csv(file_path, dtype=str)
        return name, df, None, "csv", "data"
    xls = pd.ExcelFile(file_path)
    sheet = sheet_name or (xls.sheet_names[0] if xls.sheet_names else "data")
    df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
    return name, df, xls, "excel", sheet


def run_geocode_job(
    input_path: str,
    zip_cols: List[str],
    addr_cols: List[str],
    *,
    batch_size: int = BATCH_SIZE_DEFAULT,
    uploaded_cache: Optional[Dict[str, Tuple[Optional[float], Optional[float], str]]] = None,
    sheet_name: Optional[str] = None,
    progress_cb=None,
    job_id: Optional[str] = None,
) -> Dict[str, str]:
    """
    バッチジオコーディングを実行し、出力ファイルとキャッシュファイルのパスを返す。
    progress_cb が渡された場合は progress_cb(done, total, phase, message) で進捗通知する。
    """
    base_dir = os.path.dirname(input_path)
    name, df_input, xls, kind, sheet_name = _load_input_file(input_path, sheet_name=sheet_name)
    base_name = os.path.splitext(name)[0]

    if progress_cb:
        progress_cb(0, 1, "prepare", "マスタ読込")
    master_df = logic.read_master()
    if progress_cb:
        progress_cb(0, 1, "prepare", f"マスタ読込完了 shape={master_df.shape}")

    df_work = df_input.copy()
    used_zip_codes = set()
    used_master_idx = set()

    # 郵便番号突合
    if zip_cols:
        def zip_prog(done, total, detail):
            if progress_cb:
                progress_cb(done, total, "zip", detail)

        df_work = logic.attach_master_by_zip(
            df_work, master_df, zip_cols, progress=zip_prog, used_zip_codes=used_zip_codes
        )

    # 住所突合
    if addr_cols:
        def addr_prog(col, done, total, detail):
            if progress_cb:
                progress_cb(done, total, "addr", detail)

        df_work = logic.attach_master_by_address(
            df_work, master_df, addr_cols, progress=addr_prog, used_master_idx=used_master_idx
        )

    # キャッシュ
    os.makedirs(CACHE_DIR, exist_ok=True)
    cache_file_name = f"streamlit_local_cache.json" if job_id is None else f"{job_id}_cache.json"
    local_cache_path = os.path.join(CACHE_DIR, cache_file_name)
    cache = logic.load_cache(local_cache_path)
    if uploaded_cache:
        cache.update(uploaded_cache)

    # ジオコーディング（ユニーク住所をバッチ処理）
    geo_results = {}
    if addr_cols:
        all_addrs = []
        for col in addr_cols:
            all_addrs.extend(df_work[col].dropna().tolist())
        unique_addrs = [a for a in pd.Series(all_addrs).dropna().unique().tolist() if logic.normalize_address(a)]
        total_unique = len(unique_addrs)
        overall_done = 0
        for start in range(0, total_unique, batch_size):
            end = min(start + batch_size, total_unique)
            chunk = unique_addrs[start:end]

            def geo_prog(done, total, kind):
                if progress_cb:
                    progress_cb(overall_done + done, total_unique, "geo", f"{overall_done + done}/{total_unique}")

            def geo_cache_save(c):
                logic.save_cache(local_cache_path, c)

            if progress_cb:
                progress_cb(overall_done, total_unique, "geo", f"バッチ {start+1}〜{end}")
            chunk_results, cache_hit, new_count = logic.geocode_addresses(
                chunk,
                user_agent="GeoGUI_streamlit",
                cache=cache,
                progress_cb=geo_prog,
                cache_save_cb=geo_cache_save,
            )
            geo_results.update(chunk_results)
            logic.save_cache(local_cache_path, cache)
            overall_done = end
            if progress_cb:
                progress_cb(overall_done, total_unique, "geo", f"完了 cache_hit={cache_hit} 新規={new_count}")
        df_work = logic.add_geocode_columns(df_work, addr_cols, geo_results)

    # 出力生成
    out_base = base_name or "output"
    if kind == "excel" and xls is not None:
        buf = _build_excel_output(
            xls, sheet_name, df_input, df_work, master_df, used_zip_codes, used_master_idx
        )
        fname = f"{out_base}{OUTPUT_SUFFIX}.xlsx"
    else:
        buf = _build_csv_output(df_work)
        fname = f"{out_base}{OUTPUT_SUFFIX}.csv"

    output_path = os.path.join(base_dir, fname)
    with open(output_path, "wb") as f:
        f.write(buf.getvalue())

    if progress_cb:
        progress_cb(1, 1, "done", "完了")

    return {
        "output_path": output_path,
        "cache_path": local_cache_path,
        "output_name": fname,
    }
