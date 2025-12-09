# -*- coding: utf-8 -*-
"""
このツールの使い方（要約）
1. 入力ファイルを選択（Excelならシート指定、CSVも可）
2. 郵便番号列数・住所列数を設定し、対象列をドロップダウンで選択
3. マスタ突合
   - マスタ: G:\\共有ドライブ\\03_データサイエンス事業部\\開発ネタ\\Geoコーディング\\01_Data\\02_processd\\zipcode_localgoverment_mst.xlsx
   - 郵便番号→住所の順で突合し、`<入力列名>_<マスタ列名>` を付与
   - 郵便番号突合完了で zip 用CSVチェックポイント保存、住所突合完了で addr 用CSVチェックポイント保存
   - 住所突合では「郵便番号・事業所郵便番号フラグ・大口事業所名・小字名、丁目、番地等（漢字）」は付与しない（郵便番号突合のみ付与）
4. ジオコーディング
   - Nominatimで段階検索（全文→丁目→町/村→市/区→都・道・府・県）
   - 住所はユニーク化して呼び出し、0.2秒間隔・10万件ごと7秒休止
   - 10,000件ごとにキャッシュJSONを保存（途中停止でも再開可）、完了時にジオコーディングCSVチェックポイント保存
5. 出力
   - CSV: `<元ファイル名>_GSI.csv`
   - Excel: 全シートコピー＋元シートそのまま＋付与後シート `<元シート名>_GSI`＋マスタシート
   - シート数を進捗に反映（何枚中何枚コピー）
   - マスタシートは突合で使用した行を黄色で塗りつぶし
6. 地図出力
   - foliumでユニーク住所を地図にプロットし、`<元ファイル名>_GSI_map.html` を保存
7. 進捗表示（バーとログ）
   - 郵便番号突合: 列ごとに更新
   - 住所突合: 列ごとに更新（郵便番号突合と同様のタイミング）
   - ジオコーディング: 随時更新（ユニーク住所ベース）
   - 出力: Excelはシートコピーごと、CSVは開始/完了で更新
   - ログはタイムスタンプ付き。例: `[12:00:00] 住所突合 5000/12000 (41.7%)`
"""
# 標準ライブラリ
import os
import time
import math
import json
import queue
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from typing import Dict, List, Optional, Tuple

# 外部ライブラリ
import pandas as pd
import requests
import folium
from openpyxl.styles import PatternFill  # Excelセルの塗りつぶし用

# 定数定義
MASTER_PATH = r"G:\共有ドライブ\03_データサイエンス事業部\開発ネタ\Geoコーディング\01_Data\02_processd\zipcode_localgoverment_mst.xlsx" # マスタファイルパス
OUTPUT_SUFFIX = "_GSI" # 出力ファイル名サフィックス
CHUNK_SIZE = 100_000 # Nominatim負荷を考慮
CHUNK_SLEEP_SEC = 7.0 # Nominatim負荷を考慮
REQUEST_SLEEP_SEC = 0.2  # Nominatim負荷を考慮しつつ高速化
PROGRESS_UPDATE_SEC = 60  # 進捗ログ更新間隔（秒）
CHECKPOINT_SAVE_EVERY = 10_000  # ジオコーディング時のキャッシュ保存間隔（ユニーク住所数ベース）
CACHE_DIR = os.path.join(os.path.dirname(__file__), "cache")  # 共有キャッシュ保存先
GLOBAL_CACHE_PATH = os.path.join(CACHE_DIR, "geocode_cache_global.json")  # 履歴を貯める共通辞書

# 地方コード・地方名マッピング（都道府県コード先頭2桁で判定）
REGION_MAP = {
    "01": ("1", "北海道"),
    "02": ("2", "東北"), "03": ("2", "東北"), "04": ("2", "東北"), "05": ("2", "東北"), "06": ("2", "東北"), "07": ("2", "東北"),
    "08": ("3", "関東"), "09": ("3", "関東"), "10": ("3", "関東"), "11": ("3", "関東"), "12": ("3", "関東"), "13": ("3", "関東"), "14": ("3", "関東"),
    "15": ("4", "中部"), "16": ("4", "中部"), "17": ("4", "中部"), "18": ("4", "中部"), "19": ("4", "中部"), "20": ("4", "中部"), "21": ("4", "中部"), "22": ("4", "中部"), "23": ("4", "中部"),
    "24": ("5", "近畿"), "25": ("5", "近畿"), "26": ("5", "近畿"), "27": ("5", "近畿"), "28": ("5", "近畿"), "29": ("5", "近畿"), "30": ("5", "近畿"),
    "31": ("6", "中国"), "32": ("6", "中国"), "33": ("6", "中国"), "34": ("6", "中国"), "35": ("6", "中国"),
    "36": ("7", "四国"), "37": ("7", "四国"), "38": ("7", "四国"), "39": ("7", "四国"),
    "40": ("8", "九州・沖縄"), "41": ("8", "九州・沖縄"), "42": ("8", "九州・沖縄"), "43": ("8", "九州・沖縄"),
    "44": ("8", "九州・沖縄"), "45": ("8", "九州・沖縄"), "46": ("8", "九州・沖縄"), "47": ("8", "九州・沖縄"),
}

# マスタの列名一覧（郵便番号突合で全列付与）
MASTER_COLUMNS_ZIP = [
    "郵便番号",
    "地方コード",
    "地方名",
    "都道府県コード",
    "都道府県名(漢字)",
    "団体コード",
    "市区町村名(漢字)",
    "政令指定都市フラグ",
    "町域名(漢字)",
    "小字名、丁目、番地等（漢字）",
    "事業所郵便番号フラグ",
    "大口事業所名（漢字）",
]
# 住所突合では郵便番号・事業所系・「小字名、丁目、番地等（漢字）」を除外して付与
MASTER_COLUMNS_ADDR = [
    "地方コード",
    "地方名",
    "都道府県コード",
    "都道府県名(漢字)",
    "団体コード",
    "市区町村名(漢字)",
    "政令指定都市フラグ",
    "町域名(漢字)",
]


# 空白やNaNを扱いやすい形にするユーティリティ
def safe_strip(val: Optional[str]) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)): # NaNチェック
        return "" # 空文字返却
    return str(val).strip() # 文字列化して前後空白除去

# 郵便番号を7桁ゼロ埋め
def pad_zip(val: Optional[str]) -> str:
    v = safe_strip(val)
    if v.isdigit():
        return v.zfill(7)
    return v


# マスタExcelを読み込む
def read_master() -> pd.DataFrame:
    df = pd.read_excel(MASTER_PATH, dtype=str) # マスタ読み込み
    df.columns = [c.strip() for c in df.columns] # 列名の前後空白除去
    # 地方コード・地方名を追加
    def to_region(code: str):
        if not code or not isinstance(code, str):
            return None, None
        pref2 = code[:2]
        rc, rn = REGION_MAP.get(pref2, (None, None))
        if rc is not None:
            rc = str(rc).zfill(2)  # 2桁ゼロ埋め
        return rc, rn
    region_codes = []
    region_names = []
    for code in df["都道府県コード"]:
        r_code, r_name = to_region(code)
        region_codes.append(r_code)
        region_names.append(r_name)
    df["地方コード"] = region_codes
    df["地方名"] = region_names
    return df # マスタデータフレーム返却


# 住所文字列の正規化（空白除去など）
def normalize_address(addr: str) -> str:
    if not addr: # Noneまたは空文字
        return "" # 空文字返却
    # 全角スペース・改行・タブ・半角スペースを除去して比較用に使う
    return str(addr).replace("\u3000", "").replace("\n", "").replace("\t", "").replace(" ", "").strip()


# 住所から都道府県名を抽出（部分一致、空白除去後）
def find_prefecture(addr: str, prefs: List[str]) -> Optional[str]:
    for p in prefs: # 都道府県名リストをループ
        p_norm = normalize_address(p)
        if p_norm and p_norm in addr: # 部分一致チェック
            return p # 見つかったら元の名称で返却
    return None # 見つからなければNone返却


# 住所を都道府県→市区町村→町域の順でマスタと突合し、マッチした粒度だけ列を埋める
# 判定順: (1) 完全一致（都道府県+市区町村+町域の正規化文字列）→(2) 部分一致（市区町村+町域が含まれる）→(3) 市区町村のみ含まれる→(4) 都道府県のみ
def match_master_address(addr: str, master_by_pref: Dict[str, pd.DataFrame], city_groups: Optional[Dict[str, pd.DataFrame]] = None) -> Optional[Tuple[Dict[str, Optional[str]], Optional[int], Optional[str]]]:
    """
    住所を都道府県→市区町村→町域の順で突合し、都道府県が無い場合は市区町村が一意なら補完。
    戻り値: (値辞書, 使用マスタindex, マッチフラグ)
    """
    addr_norm = normalize_address(addr)
    if not addr_norm:
        return None

    result: Dict[str, Optional[str]] = {col: None for col in MASTER_COLUMNS_ADDR}
    idx_used: Optional[int] = None
    match_flag: Optional[str] = None

    pref = find_prefecture(addr_norm, list(master_by_pref.keys()))
    if pref:
        df_pref = master_by_pref.get(pref)
        if df_pref is None or df_pref.empty:
            return None

        base_row = df_pref.iloc[0]
        result["地方コード"] = base_row.get("地方コード")
        result["地方名"] = base_row.get("地方名")
        result["都道府県コード"] = base_row["都道府県コード"]
        result["都道府県名(漢字)"] = base_row["都道府県名(漢字)"]

        df_pref_sorted = df_pref.copy()
        df_pref_sorted["len_city_town"] = df_pref_sorted["市区町村名(漢字)"].fillna("").str.len() + df_pref_sorted["町域名(漢字)"].fillna("").str.len()
        df_pref_sorted = df_pref_sorted.sort_values("len_city_town", ascending=False)

        # 1. 完全一致（pref+city+town が先頭一致）
        for _, row in df_pref_sorted.iterrows():
            city = safe_strip(row["市区町村名(漢字)"])
            town = safe_strip(row["町域名(漢字)"])
            full_norm = normalize_address(f"{pref}{city}{town}")
            if full_norm and addr_norm.startswith(full_norm):
                result.update({
                    "地方コード": row.get("地方コード", result.get("地方コード")),
                    "地方名": row.get("地方名", result.get("地方名")),
                    "団体コード": row["団体コード"],
                    "市区町村名(漢字)": row["市区町村名(漢字)"],
                    "政令指定都市フラグ": row["政令指定都市フラグ"],
                    "町域名(漢字)": row["町域名(漢字)"],
                })
                idx_used = row.name
                match_flag = "pref_city_town"
                return result, idx_used, match_flag

        # 2. 市区町村＋町域の部分一致
        for _, row in df_pref_sorted.iterrows():
            city = safe_strip(row["市区町村名(漢字)"])
            town = safe_strip(row["町域名(漢字)"])
            city_norm = normalize_address(city)
            town_norm = normalize_address(town)
            if city_norm and city_norm in addr_norm and town_norm and town_norm in addr_norm:
                result.update({
                    "地方コード": row.get("地方コード", result.get("地方コード")),
                    "地方名": row.get("地方名", result.get("地方名")),
                    "団体コード": row["団体コード"],
                    "市区町村名(漢字)": row["市区町村名(漢字)"],
                    "政令指定都市フラグ": row["政令指定都市フラグ"],
                    "町域名(漢字)": row["町域名(漢字)"],
                })
                idx_used = row.name
                match_flag = "pref_city_town_partial"
                return result, idx_used, match_flag

        # 3. 市区町村のみ
        for _, row in df_pref.iterrows():
            city = safe_strip(row["市区町村名(漢字)"])
            city_norm = normalize_address(city)
            if city_norm and city_norm in addr_norm:
                result.update({
                    "地方コード": row.get("地方コード", result.get("地方コード")),
                    "地方名": row.get("地方名", result.get("地方名")),
                    "団体コード": row["団体コード"],
                    "市区町村名(漢字)": row["市区町村名(漢字)"],
                    "政令指定都市フラグ": row["政令指定都市フラグ"],
                    "町域名(漢字)": None,
                })
                idx_used = row.name
                match_flag = "pref_city"
                return result, idx_used, match_flag

        match_flag = "pref_only"
        return result, idx_used, match_flag

    # 都道府県が無い場合: 市区町村が住所に含まれ一意なら補完
    if city_groups:
        matched = []
        for city_norm, df_city in city_groups.items():
            if city_norm and city_norm in addr_norm:
                matched.append(df_city)
        if len(matched) == 1:
            df_city = matched[0].copy()
            df_city["len_town"] = df_city["町域名(漢字)"].fillna("").str.len()
            df_city = df_city.sort_values("len_town", ascending=False)
            for _, row in df_city.iterrows():
                town = safe_strip(row["町域名(漢字)"])
                town_norm = normalize_address(town)
                if town_norm and town_norm in addr_norm:
                    result.update({
                        "地方コード": row.get("地方コード"),
                        "地方名": row.get("地方名"),
                        "都道府県コード": row["都道府県コード"],
                        "都道府県名(漢字)": row["都道府県名(漢字)"],
                        "団体コード": row["団体コード"],
                        "市区町村名(漢字)": row["市区町村名(漢字)"],
                        "政令指定都市フラグ": row["政令指定都市フラグ"],
                        "町域名(漢字)": row["町域名(漢字)"],
                    })
                    idx_used = row.name
                    match_flag = "no_pref_city_town"
                    return result, idx_used, match_flag
            # 町域は不明だが市区町村が一意
            row = df_city.iloc[0]
            result.update({
                "地方コード": row.get("地方コード"),
                "地方名": row.get("地方名"),
                "都道府県コード": row["都道府県コード"],
                "都道府県名(漢字)": row["都道府県名(漢字)"],
                "団体コード": row["団体コード"],
                "市区町村名(漢字)": row["市区町村名(漢字)"],
                "政令指定都市フラグ": row["政令指定都市フラグ"],
                "町域名(漢字)": None,
            })
            idx_used = row.name
            match_flag = "no_pref_city"
            return result, idx_used, match_flag

    return None


# Nominatim APIを1回呼び出す
def nominatim_search(query: str, user_agent: str) -> Optional[Tuple[float, float]]:
    url = "https://nominatim.openstreetmap.org/search" # Nominatim検索URL
    params = {"q": query, "format": "json", "limit": 1}
    headers = {"User-Agent": user_agent}
    try:
        res = requests.get(url, params=params, headers=headers, timeout=10)
        if res.status_code == 200: # 正常応答
            data = res.json()
            if isinstance(data, list) and data: # 結果あり
                return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception:  # タイムアウトなどはNoneで返す
        return None
    return None


# 段階的に狭めた検索クエリを生成
def generate_queries(addr: str) -> List[Tuple[str, str]]:
    addr = normalize_address(addr)
    queries = []
    if addr:
        queries.append((addr, "full"))
        if "丁目" in addr:  # 丁目でカット（フラグ名もtokenと合わせる）
            queries.append((addr.split("丁目")[0] + "丁目", "town"))
        for token in ["町", "村"]:  # 町・村でカット
            if token in addr:
                queries.append((addr.split(token)[0] + token, "town"))
        for token in ["市", "区"]:  # 市・区でカット
            if token in addr:
                queries.append((addr.split(token)[0] + token, "city"))
        if "県" in addr:  # 都道府県レベルまで縮約
            queries.append((addr.split("県")[0] + "県", "pref"))
        elif "都" in addr:
            queries.append((addr.split("都")[0] + "都", "pref"))
        elif "府" in addr:
            queries.append((addr.split("府")[0] + "府", "pref"))
    return queries


# JSONキャッシュを読み込む（存在しなければ空dict）
def load_cache(cache_path: str) -> Dict[str, Tuple[Optional[float], Optional[float], str]]:
    """JSONキャッシュを読み込む（存在しなければ空dict）"""
    if not cache_path or not os.path.exists(cache_path):
        return {}
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        # 値は [lat, lon, flag] のリストとして保存、タプルに戻す
        return {k: (v[0], v[1], v[2]) for k, v in data.items()}
    except Exception:  # 壊れたファイルなどは無視して空で返す
        return {}


# JSONキャッシュを書き出す
def save_cache(cache_path: str, cache: Dict[str, Tuple[Optional[float], Optional[float], str]]):
    """JSONキャッシュを書き出す"""
    if not cache_path:
        return
    try:
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        serializable = {k: [v[0], v[1], v[2]] for k, v in cache.items()} # タプルをリストに変換して保存
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(serializable, f, ensure_ascii=False)
    except Exception:  # 書き込み失敗は握りつぶす（致命的ではないため）
        pass


# 住所をユニーク化し、キャッシュを利用しながらジオコーディングを実施
def geocode_addresses(addresses: List[str], user_agent: str, cache: Dict[str, Tuple[Optional[float], Optional[float], str]], progress_cb=None, cache_save_cb=None) -> Tuple[Dict[str, Tuple[Optional[float], Optional[float], str]], int, int]:
    """住所をユニーク化し、キャッシュを利用しながらジオコーディングを実施"""
    results: Dict[str, Tuple[Optional[float], Optional[float], str]] = {} # 住所→(lat, lon, flag)
    unique_addrs = [a for a in pd.Series(addresses).dropna().unique().tolist() if normalize_address(a)] # 空文字除去
    total = len(unique_addrs)
    cache_hit = 0
    new_count = 0
    if progress_cb:
        progress_cb(0, total, "start")
    last_log = time.time()
    for i, addr in enumerate(unique_addrs, start=1):
        norm = normalize_address(addr)
        if norm in cache:
            results[addr] = cache[norm]
            cache_hit += 1
        else:
            found = False
            for query, flag in generate_queries(addr):  # 段階検索
                res = nominatim_search(query, user_agent)
                if res:
                    results[addr] = (res[0], res[1], flag)
                    cache[norm] = (res[0], res[1], flag)
                    found = True
                    new_count += 1
                    break
                time.sleep(REQUEST_SLEEP_SEC)
            if not found:  # すべて失敗した場合
                results[addr] = (None, None, "not_found")
                cache[norm] = (None, None, "not_found")
                new_count += 1
        # 一定件数ごとにキャッシュを保存（耐障害性向上）
        if cache_save_cb and (i % CHECKPOINT_SAVE_EVERY == 0):
            cache_save_cb(cache)
        if i % CHUNK_SIZE == 0:
            time.sleep(CHUNK_SLEEP_SEC)
        if progress_cb:
            now = time.time()
            if now - last_log >= PROGRESS_UPDATE_SEC:
                progress_cb(i, total, "tick")
                last_log = now
    if progress_cb:
        progress_cb(total, total, "done")
    return results, cache_hit, new_count


# 郵便番号列でマスタ突合
def attach_master_by_zip(df: pd.DataFrame, master: pd.DataFrame, zip_cols: List[str], progress=None, used_zip_codes=None) -> pd.DataFrame:
    result = df.copy()
    total = max(len(zip_cols), 1)
    done = 0
    if used_zip_codes is None:
        used_zip_codes = set()
    for col in zip_cols:
        if col not in result.columns:
            done += 1
            if progress:
                progress(done, total, f"{col} 列なし")
            continue
        tmp = result[[col]].rename(columns={col: "郵便番号"})
        tmp["郵便番号"] = tmp["郵便番号"].apply(pad_zip)
        master_zip = master.copy()
        master_zip["郵便番号"] = master_zip["郵便番号"].apply(pad_zip)
        merged = pd.merge(tmp, master_zip, on="郵便番号", how="left")
        for mcol in MASTER_COLUMNS_ZIP:
            new_col = f"{col}_{mcol}"
            if new_col not in result.columns:
                result[new_col] = merged[mcol]
            else:
                result[new_col] = result[new_col].fillna(merged[mcol])
        # 使用した郵便番号を記録
        matched_zips = merged["郵便番号"].dropna().unique().tolist()
        used_zip_codes.update(matched_zips)
        done += 1
        if progress:
            progress(done, total, f"{col} 突合完了")
    return result


# 住所列でマスタ突合
def attach_master_by_address(df: pd.DataFrame, master: pd.DataFrame, addr_cols: List[str], progress=None, used_master_idx=None) -> pd.DataFrame:
    if not addr_cols:
        return df.copy()
    result = df.copy()
    # 住所突合は事業所郵便番号フラグが'0'（対象外）だけを使う
    master_addr = master[master["事業所郵便番号フラグ"] == "0"].copy()
    master_by_pref: Dict[str, pd.DataFrame] = {p: g for p, g in master_addr.groupby("都道府県名(漢字)")}
    # 都道府県なし用: 市区町村正規化ごとにグループを持つ
    master_addr["city_norm"] = master_addr["市区町村名(漢字)"].fillna("").apply(normalize_address)
    city_groups: Dict[str, pd.DataFrame] = {k: v for k, v in master_addr.groupby("city_norm")}
    cache: Dict[str, Optional[pd.Series]] = {}
    if used_master_idx is None:
        used_master_idx = set()
    total_cols = max(len(addr_cols), 1)
    done_cols = 0
    for col in addr_cols:
        if col not in result.columns:
            continue
        # ユニーク住所で突合
        unique_addrs = pd.Series(result[col].fillna("")).unique().tolist()
        for addr in unique_addrs:
            addr_norm = normalize_address(addr)
            if addr_norm not in cache:
                matched = match_master_address(addr_norm, master_by_pref, city_groups=city_groups)
                if matched is not None:
                    vals_dict, idx_val, flag = matched
                    cache[addr_norm] = (vals_dict, idx_val, flag)
                    if idx_val is not None:
                        used_master_idx.add(idx_val)
                else:
                    cache[addr_norm] = (None, None, None)
        # 行に結果を展開
        records = []
        for addr in result[col].fillna(""):
            addr_norm = normalize_address(addr)
            records.append(cache.get(addr_norm))
        # 住所突合では事業所系と「小字名、丁目、番地等（漢字）」は付与しない（郵便番号突合のみ付与）
        for mcol in MASTER_COLUMNS_ADDR:
            new_col = f"{col}_{mcol}"
            vals = []
            for rec in records:
                if rec is None or rec[0] is None:
                    vals.append(None)
                else:
                    vals.append(rec[0].get(mcol, None))
            result[new_col] = vals
        # マッチフラグ列を追加
        flag_col = f"{col}_match_flag"
        flags = []
        for rec in records:
            if rec is None:
                flags.append(None)
            else:
                flags.append(rec[2])
        result[flag_col] = flags
        done_cols += 1
        if progress:
            progress(col, done_cols, total_cols, f"{col} 突合完了")
    return result


# 緯度経度・フラグ列をDataFrameへ埋め込む
def add_geocode_columns(df: pd.DataFrame, addr_cols: List[str], results: Dict[str, Tuple[Optional[float], Optional[float], str]]) -> pd.DataFrame:
    out = df.copy()
    # フラグ表記を英語にそろえる（キャッシュに旧表記が残っている場合の補正）
    def normalize_flag(flag: Optional[str]) -> Optional[str]:
        mapping = {
            "完全一致": "full",
            "full": "full",
            "丁目": "town",
            "町": "town",
            "村": "town",
            "town": "town",
            "市": "city",
            "区": "city",
            "city": "city",
            "pref": "pref",
            "not_found": "not_found",
            None: None,
        }
        return mapping.get(flag, flag)  # 未知の表記はそのまま
    for col in addr_cols:
        lat_col = f"{col}_lat"
        lon_col = f"{col}_lon"
        flag_col = f"{col}_geocode_flag"
        if lat_col not in out.columns:  # 足りない列は新規作成
            out[lat_col] = None
        if lon_col not in out.columns:
            out[lon_col] = None
        if flag_col not in out.columns:
            out[flag_col] = None
        for idx, addr in out[col].items():
            res = results.get(addr)
            if res:  # 結果があるときだけ上書き
                out.at[idx, lat_col] = res[0]
                out.at[idx, lon_col] = res[1]
                out.at[idx, flag_col] = normalize_flag(res[2])
    return out


# foliumで地図HTMLを生成
def create_map_html(df: pd.DataFrame, addr_cols: List[str], output_path: str):
    points = []
    for col in addr_cols:
        lat_col = f"{col}_lat"
        lon_col = f"{col}_lon"
        if lat_col not in df.columns or lon_col not in df.columns:
            continue
        subset = df[[col, lat_col, lon_col]].dropna()
        for _, row in subset.iterrows():
            try:
                lat = float(row[lat_col])
                lon = float(row[lon_col])
            except Exception:
                continue
            if math.isnan(lat) or math.isnan(lon):
                continue
            points.append((row[col], lat, lon))
    if not points:
        return None
    # 地図中央を平均座標に設定
    avg_lat = sum(p[1] for p in points) / len(points)
    avg_lon = sum(p[2] for p in points) / len(points)
    fmap = folium.Map(location=[avg_lat, avg_lon], zoom_start=8)
    # マーカーを追加
    for addr, lat, lon in points:
        folium.Marker([lat, lon], tooltip=addr).add_to(fmap)
    fmap.save(output_path)
    return output_path


# Excelで出力（全シートコピー＋処理済シート＋マスタ）
def write_output_excel(input_path: str, output_dir: str, sheet_name: str, original_df: pd.DataFrame, processed_df: pd.DataFrame, master: pd.DataFrame, log_func=None, progress_cb=None, highlight_idx=None) -> str:
    base = os.path.splitext(os.path.basename(input_path))[0]
    out_path = os.path.join(output_dir, f"{base}{OUTPUT_SUFFIX}.xlsx")
    all_sheets = pd.read_excel(input_path, sheet_name=None, dtype=str) # 全シート読み込み
    total_sheets = len(all_sheets)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for idx, (name, df) in enumerate(all_sheets.items(), start=1):
            if log_func:
                log_func(f"シートコピー開始: {name}")
            df.to_excel(writer, sheet_name=name, index=False)
            if log_func:
                log_func(f"シートコピー完了: {name}")
            # シートコピー進捗を更新
            if progress_cb:
                progress_cb(idx, total_sheets, f"シートコピー")
        original_df.to_excel(writer, sheet_name=sheet_name, index=False)
        processed_df.to_excel(writer, sheet_name=f"{sheet_name}{OUTPUT_SUFFIX}", index=False)
        master.to_excel(writer, sheet_name="master", index=False)
        # 使用行を黄色塗りつぶし
        if highlight_idx:
            ws = writer.book["master"]
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for idx in highlight_idx:
                row_num = idx + 2  # pandasのindexは0始まり、Excelは1始まり＋ヘッダー1行
                for col_num in range(1, len(master.columns) + 1):
                    ws.cell(row=row_num, column=col_num).fill = fill
    if progress_cb:
        progress_cb(total_sheets, total_sheets, "シートコピー完了")
    return out_path


# CSVで出力
def write_output_csv(input_path: str, output_dir: str, processed_df: pd.DataFrame, log_func=None) -> str:
    base = os.path.splitext(os.path.basename(input_path))[0]
    out_path = os.path.join(output_dir, f"{base}{OUTPUT_SUFFIX}.csv")
    if log_func:
        log_func("CSV出力開始")
    processed_df.to_csv(out_path, index=False, encoding="utf-8-sig")
    if log_func:
        log_func("CSV出力完了")
    return out_path


class GeoApp:
    def __init__(self, root: tk.Tk):
        # GUIと処理全体のハンドラ
        self.root = root
        self.root.title("Geo付与ツール") # ウィンドウタイトル
        self.file_path: Optional[str] = None # 入力ファイルパス
        self.sheet_name: Optional[str] = None # 選択シート名
        self.df: Optional[pd.DataFrame] = None # 入力データフレーム
        self.master_df = read_master() # マスタデータフレーム
        self.master_by_pref = {p: g for p, g in self.master_df.groupby("都道府県名(漢字)")} # マスタ都道府県別辞書
        self.output_dir: Optional[str] = None # 出力フォルダパス
        self.zip_count_var = tk.IntVar(value=1) # 郵便番号列数
        self.addr_count_var = tk.IntVar(value=1) # 住所列数
        self.zip_selectors: List[ttk.Combobox] = [] # 郵便番号列選択コンボボックス
        self.addr_selectors: List[ttk.Combobox] = [] # 住所列選択コンボボックス
        self.status_var = tk.StringVar(value="ファイルを選択してください") # ステータス表示用
        self.progress_var = tk.DoubleVar(value=0) # 進捗バー用
        self.progress_total = 1 # 進捗総数
        self.log_text: Optional[tk.Text] = None # ログ表示用テキストウィジェット
        # 進捗共有用の状態（1分ごとにタイマーが参照）
        self.progress_state = {"phase": "idle", "done": 0, "total": 1, "detail": ""} # 進捗状態辞書
        self.progress_running = False # 進捗タイマー稼働フラグ
        self.progress_thread: Optional[threading.Thread] = None # 進捗タイマースレッド
        self.used_zip_codes = set()  # 郵便番号突合でマッチした郵便番号を記録（マスタ着色用）
        self.used_master_idx = set()  # 住所突合でマッチしたマスタ行indexを記録（マスタ着色用）
        self._build_ui() # UI構築

    def _build_ui(self):
        # GUIパーツの配置
        frm = tk.Frame(self.root) # メインフレーム
        frm.pack(fill=tk.BOTH, expand=True, padx=10, pady=10) # フレーム配置
        frm.grid_columnconfigure(0, weight=1) # 列0を伸縮可能に
        frm.grid_columnconfigure(1, weight=1) # 列1を伸縮可能に

        # 入力ファイル選択
        tk.Button(frm, text="入力ファイル選択", command=self.select_file).grid(row=0, column=0, sticky="w")
        self.file_label = tk.Label(frm, text="未選択", anchor="w")
        self.file_label.grid(row=0, column=1, sticky="w")

        #  Excelシート選択
        tk.Label(frm, text="Excelシート").grid(row=1, column=0, sticky="w")
        self.sheet_var = tk.StringVar()
        self.sheet_menu = ttk.Combobox(frm, textvariable=self.sheet_var, state="disabled")
        self.sheet_menu.grid(row=1, column=1, sticky="w")
        self.sheet_menu.bind("<<ComboboxSelected>>", lambda e: self.load_selected_sheet())

        # 郵便番号/住所列数選択
        tk.Label(frm, text="郵便番号列数").grid(row=2, column=0, sticky="w")
        tk.Spinbox(frm, from_=0, to=5, textvariable=self.zip_count_var, width=5, command=self.update_selectors).grid(row=2, column=1, sticky="w")

        # 住所列数選択
        tk.Label(frm, text="住所列数").grid(row=3, column=0, sticky="w")
        tk.Spinbox(frm, from_=0, to=5, textvariable=self.addr_count_var, width=5, command=self.update_selectors).grid(row=3, column=1, sticky="w")

        # 郵便番号/住所列選択フレーム
        self.zip_frame = tk.Frame(frm) # 郵便番号列選択フレーム
        self.zip_frame.grid(row=4, column=0, columnspan=2, sticky="we") # 郵便番号列選択フレーム
        self.zip_frame.grid_columnconfigure(0, weight=0) # ラベル列は伸縮不可
        self.zip_frame.grid_columnconfigure(1, weight=1) # コンボ列は伸縮可能
        self.addr_frame = tk.Frame(frm) # 住所列選択フレーム
        self.addr_frame.grid(row=5, column=0, columnspan=2, sticky="we") # 住所列選択フレーム
        self.addr_frame.grid_columnconfigure(0, weight=0) # ラベル列は伸縮不可
        self.addr_frame.grid_columnconfigure(1, weight=1) # コンボ列は伸縮可能

        # 出力フォルダ選択
        tk.Button(frm, text="出力フォルダ選択", command=self.select_output_dir).grid(row=6, column=0, sticky="w")
        self.output_label = tk.Label(frm, text="未選択", anchor="w")
        self.output_label.grid(row=6, column=1, sticky="w")
        
        # 実行ボタン
        self.run_button = tk.Button(frm, text="実行", command=self.run, state="disabled")
        # ボタンは中央寄せ、幅はデフォルトサイズ
        self.run_button.grid(row=7, column=0, columnspan=2, pady=5) # 実行ボタン配置

        # 進捗バーとステータス表示
        ttk.Progressbar(frm, variable=self.progress_var, maximum=100).grid(row=8, column=0, columnspan=2, sticky="we", padx=5) # 進捗バー配置
        self.status_label = tk.Label(frm, textvariable=self.status_var, anchor="w") # ステータスラベル
        self.status_label.grid(row=9, column=0, columnspan=2, sticky="w") # ステータスラベル配置

        # ログ表示用テキスト（フレーム配下でgrid管理）
        self.log_text = tk.Text(frm, height=8, width=100, state="disabled") # ログ表示用テキストウィジェット
        self.log_text.grid(row=10, column=0, columnspan=2, sticky="we", pady=5) # テキストウィジェット配置

    # ログ追記用メソッド
    def log(self, msg: str):
        """UIテキストにログを追記（タイムスタンプ付与）"""
        timestamp = time.strftime("%H:%M:%S") # 時刻のみ表示
        line = f"[{timestamp}] {msg}" # ログ行生成
        self.log_text.configure(state="normal") # 書き込み可能に
        self.log_text.insert(tk.END, line + "\n") # 行追加
        self.log_text.see(tk.END) # 最下行にスクロール
        self.log_text.configure(state="disabled") # 再び書き込み不可に

    # 進捗タイマー関連メソッド
    def start_progress_timer(self):
        """1分ごとに現在のフェーズと進捗をログするバックグラウンドタイマーを起動"""
        if self.progress_running:
            return
        self.progress_running = True
        self.progress_thread = threading.Thread(target=self._progress_ticker, daemon=True) # デーモンスレッド
        self.progress_thread.start()

    # タイマー停止メソッド
    def stop_progress_timer(self):
        """タイマー停止"""
        self.progress_running = False
        if self.progress_thread:
            self.progress_thread.join(timeout=1)
            self.progress_thread = None

    # 進捗更新コールバック
    def _progress_ticker(self):
        while self.progress_running:
            st = self.progress_state
            pct = 0
            if st["total"]:
                pct = (st["done"] / st["total"]) * 100 # 進捗％計算
            self.log(f"進捗[{st['phase']}]: {st['done']}/{st['total']} ({pct:.1f}%) {st['detail']}") # 進捗ログ出力
            # 1秒刻みで抜けられるように細分化
            for _ in range(PROGRESS_UPDATE_SEC):
                if not self.progress_running:
                    return
                time.sleep(1)

    def select_file(self):
        # 入力ファイルを選択し、列選択用ドロップダウンを更新
        path = filedialog.askopenfilename(title="Excel/CSVファイルを選択", filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), ("All", "*.*")])
        if not path:
            return
        self.file_path = path
        self.file_label.config(text=os.path.basename(path)) # ファイル名表示
        ext = os.path.splitext(path)[1].lower() # 拡張子取得
        if ext in [".xlsx", ".xls"]:  # Excelならシート選択を有効化
            sheets = pd.ExcelFile(path).sheet_names # シート名取得
            self.sheet_menu["values"] = sheets # シート名設定
            self.sheet_menu.config(state="readonly") # 有効化
            if sheets: # 最初のシートを選択して読み込み
                self.sheet_menu.current(0) # 最初のシートを選択
                self.sheet_name = sheets[0] # シート名設定
                self.load_selected_sheet() # シート読み込み
        else:  # CSVならシート選択は無効
            self.sheet_menu.set("") # クリア
            self.sheet_menu.config(state="disabled") # 無効化
            self.load_csv_columns() # CSV列読み込み
        self.update_run_button_state() # 実行ボタン状態更新

    def load_selected_sheet(self):
        # Excelシートを読み込み、列情報を反映
        if not self.file_path:
            return
        self.sheet_name = self.sheet_var.get()
        try:
            self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, dtype=str)
        except Exception as e:
            messagebox.showerror("読み込みエラー", str(e))
            return
        self.update_selectors()

    def load_csv_columns(self):
        # CSVを読み込み、列情報を反映
        if not self.file_path:
            return
        try:
            self.df = pd.read_csv(self.file_path, dtype=str)
        except Exception as e:
            messagebox.showerror("読み込みエラー", str(e))
            return
        self.update_selectors()

    def update_selectors(self):
        # 郵便番号/住所列のコンボボックスを個数分生成
        if self.df is None:
            return
        cols = list(self.df.columns)
        for widget in self.zip_frame.winfo_children(): # 既存ウィジェット削除
            widget.destroy()
        for widget in self.addr_frame.winfo_children(): # 既存ウィジェット削除
            widget.destroy()
        self.zip_selectors.clear()
        self.addr_selectors.clear()
        for i in range(self.zip_count_var.get()): # 郵便番号列コンボボックス生成
            tk.Label(self.zip_frame, text=f"郵便番号{i+1}", anchor="w", width=18).grid(row=i, column=0, sticky="w") # ラベル配置
            var = tk.StringVar() # 変数生成
            cb = ttk.Combobox(self.zip_frame, values=cols, textvariable=var, state="readonly", width=40) # コンボボックス生成
            cb.grid(row=i, column=1, sticky="w") # コンボボックス配置
            self.zip_selectors.append(cb) # リストに追加
        for i in range(self.addr_count_var.get()): # 住所列コンボボックス生成
            tk.Label(self.addr_frame, text=f"住所{i+1}", anchor="w", width=18).grid(row=i, column=0, sticky="w") # ラベル配置
            var = tk.StringVar() # 変数生成
            cb = ttk.Combobox(self.addr_frame, values=cols, textvariable=var, state="readonly", width=40) # コンボボックス生成
            cb.grid(row=i, column=1, sticky="w") # コンボボックス配置
            self.addr_selectors.append(cb) # リストに追加

    def select_output_dir(self):
        # 出力フォルダを選択
        directory = filedialog.askdirectory(title="出力フォルダ選択")
        if directory:
            self.output_dir = directory
            self.output_label.config(text=directory)
        self.update_run_button_state()

    def update_run_button_state(self):
        # 実行ボタンの有効/無効状態を更新
        enabled = self.file_path is not None and self.output_dir is not None
        self.run_button.config(state="normal" if enabled else "disabled") # 実行ボタン状態更新

    def run(self):
        # 実行ボタン押下時のハンドラ
        if self.df is None or not self.file_path:
            messagebox.showwarning("警告", "ファイルを選択してください")
            return
        zip_cols = [cb.get() for cb in self.zip_selectors if cb.get()]
        addr_cols = [cb.get() for cb in self.addr_selectors if cb.get()]
        if not zip_cols and not addr_cols:  # どちらも選ばれていない
            messagebox.showwarning("警告", "郵便番号または住所列を指定してください")
            return
        self.status_var.set("処理中...")
        self.run_button.config(state="disabled")
        threading.Thread(target=self._run_worker, args=(zip_cols, addr_cols), daemon=True).start()

    def _run_worker(self, zip_cols: List[str], addr_cols: List[str]):
        # バックグラウンドで実処理（マスタ突合、ジオコーディング、出力）を行う
        start_time = time.time()
        self.log("処理開始")
        self.progress_state = {"phase": "start", "done": 0, "total": 1, "detail": "開始"}
        self.start_progress_timer()
        try:
            self.progress_var.set(0)
            self.status_var.set("マスタ付与中")
            # チェックポイントパス
            base_name = os.path.splitext(os.path.basename(self.file_path))[0]
            local_dir = self.output_dir or os.path.dirname(self.file_path)
            master_zip_ckpt = os.path.join(local_dir, f"{base_name}{OUTPUT_SUFFIX}_master_zip_ckpt.csv")
            master_addr_ckpt = os.path.join(local_dir, f"{base_name}{OUTPUT_SUFFIX}_master_addr_ckpt.csv")
            geocode_ckpt = os.path.join(local_dir, f"{base_name}{OUTPUT_SUFFIX}_geocode_ckpt.csv")
            local_cache_path = os.path.join(local_dir, f"{base_name}{OUTPUT_SUFFIX}_geocode_cache.json")

            # 既存チェックポイントの確認
            if os.path.exists(geocode_ckpt):
                self.log(f"ジオコーディング済みチェックポイント読込: {geocode_ckpt}")
                df_work = pd.read_csv(geocode_ckpt, dtype=str)
                skip_zip = True
                skip_addr = True
                skip_geocode = True
            elif os.path.exists(master_addr_ckpt):
                self.log(f"住所突合チェックポイント読込: {master_addr_ckpt}")
                df_work = pd.read_csv(master_addr_ckpt, dtype=str)
                skip_zip = True
                skip_addr = True
                skip_geocode = False
            elif os.path.exists(master_zip_ckpt):
                self.log(f"郵便番号突合チェックポイント読込: {master_zip_ckpt}")
                df_work = pd.read_csv(master_zip_ckpt, dtype=str)
                skip_zip = True
                skip_addr = False
                skip_geocode = False
            else:
                df_work = self.df.copy()
                skip_zip = False
                skip_addr = False
                skip_geocode = False

            # 郵便番号突合（列ごとに進捗を更新）
            if (not skip_zip) and zip_cols:
                self.log("郵便番号突合開始")
                self.progress_state = {"phase": "master_zip", "done": 0, "total": max(len(zip_cols), 1), "detail": "郵便番号突合"}
                df_work = attach_master_by_zip(
                    df_work,
                    self.master_df,
                    zip_cols,
                    progress=lambda d, t, detail: self._on_master_progress("master_zip", d, t, detail),
                    used_zip_codes=self.used_zip_codes,
                )
                # 郵便番号突合完了チェックポイント保存
                df_work.to_csv(master_zip_ckpt, index=False, encoding="utf-8-sig")
                self.log(f"郵便番号突合チェックポイント保存: {master_zip_ckpt}")
            else:
                self.log("郵便番号突合スキップ")
            # 住所突合（ユニーク住所単位で進捗を更新）
            if (not skip_addr) and addr_cols:
                self.log("住所突合開始")
                total_addr_cols = max(len(addr_cols), 1)
                self.progress_state = {"phase": "master_addr", "done": 0, "total": total_addr_cols, "detail": "住所突合"}
                self.progress_var.set(0)
                self.status_var.set(f"住所突合 0/{total_addr_cols}")
                df_work = attach_master_by_address(df_work, self.master_df, addr_cols, progress=self._on_master_addr_progress, used_master_idx=self.used_master_idx)
                # マスタ突合完了チェックポイント保存
                df_work.to_csv(master_addr_ckpt, index=False, encoding="utf-8-sig")
                self.log(f"住所突合チェックポイント保存: {master_addr_ckpt}")
            else:
                self.log("住所突合スキップ")

            # ジオコーディング
            if skip_geocode:
                self.log("ジオコーディング済みチェックポイントを使用してスキップ")
            else:
                self.status_var.set("ジオコーディング中")
                addr_values = []
                for col in addr_cols:
                    if col in df_work.columns:
                        addr_values.extend(df_work[col].tolist())
                unique_count = len(pd.Series(addr_values).dropna().unique())
                self.log(f"ジオコーディング開始（ユニーク住所: {unique_count}件）")
                self.progress_state = {"phase": "geocode", "done": 0, "total": max(unique_count, 1), "detail": "ジオコーディング"}
                # 共通キャッシュとファイル固有キャッシュをマージして利用
                local_cache_path = os.path.join(self.output_dir or os.path.dirname(self.file_path), f"{os.path.splitext(os.path.basename(self.file_path))[0]}{OUTPUT_SUFFIX}_geocode_cache.json")
                cache_global = load_cache(GLOBAL_CACHE_PATH)
                cache_local = load_cache(local_cache_path)
                merged_cache = {**cache_global, **cache_local}
                self.log(f"キャッシュロード: グローバル{len(cache_global)}件 / ローカル{len(cache_local)}件")
                results, cache_hit, new_count = geocode_addresses(
                    addr_values,
                    user_agent="GeoGUI/1.0",
                    cache=merged_cache,
                    progress_cb=self._progress_callback,
                    cache_save_cb=lambda c: save_cache(local_cache_path, c),
                )
                save_cache(local_cache_path, merged_cache) # ファイル固有キャッシュ保存
                save_cache(GLOBAL_CACHE_PATH, merged_cache)  # 履歴を貯める共通辞書
                self.log(f"キャッシュヒット: {cache_hit}件 / 新規取得: {new_count}件 / 保存先: {local_cache_path} / グローバル更新")
                df_work = add_geocode_columns(df_work, addr_cols, results)
                # ジオコーディング完了チェックポイント保存
                df_work.to_csv(geocode_ckpt, index=False, encoding="utf-8-sig")
                self.log(f"ジオコーディングチェックポイント保存: {geocode_ckpt}")

            ext = os.path.splitext(self.file_path)[1].lower()
            if ext in [".xlsx", ".xls"]:
                all_sheets = pd.read_excel(self.file_path, sheet_name=None, dtype=str)
                sheet_name = self.sheet_name or list(all_sheets.keys())[0]
                processed_df = df_work
                original_df = all_sheets.get(sheet_name, self.df)
                self.log("Excel出力開始")
                total_steps = max(len(all_sheets), 1)  # 入力シート数で進捗管理
                self.progress_state = {"phase": "output", "done": 0, "total": total_steps, "detail": "シートコピー"}
                self.progress_var.set(0)
                self.status_var.set(f"シートコピー 0/{total_steps}")
                out_path = write_output_excel(
                    self.file_path,
                    self.output_dir,
                    sheet_name,
                    original_df,
                    processed_df,
                    self.master_df,
                    log_func=self.log,
                    progress_cb=self._on_output_progress,
                    highlight_idx=self._build_master_highlight_indices(),
                )
                # シートコピーが終わったら100%に
                self._on_output_progress(total_steps, total_steps, "シートコピー")
                self.log(f"Excel出力完了: {out_path}")
            else:
                self.log("CSV出力開始")
                self.progress_state = {"phase": "output", "done": 0, "total": 1, "detail": "CSV出力"}
                self.progress_var.set(0)
                self.status_var.set("CSV出力 0/1")
                out_path = write_output_csv(self.file_path, self.output_dir, df_work, log_func=self.log)
                self.progress_state = {"phase": "output", "done": 1, "total": 1, "detail": "CSV出力"}
                self.progress_var.set(100)
            self.status_var.set("CSV出力 1/1")
            self.log(f"CSV出力完了: {out_path}")

            map_path = os.path.join(self.output_dir, f"{os.path.splitext(os.path.basename(self.file_path))[0]}{OUTPUT_SUFFIX}_map.html")
            self.log("地図出力開始")
            map_result = create_map_html(df_work, addr_cols, map_path)
            elapsed = time.time() - start_time
            msg = f"完了: {out_path}\n" + (f"地図: {map_result}\n" if map_result else "地図: 出力なし (座標なし)\n") + f"処理時間: {elapsed:.1f}秒"
            self.status_var.set("完了")
            self.log(msg)
            messagebox.showinfo("完了", msg)
            # 完了後にチェックポイントを削除
            self._cleanup_checkpoints([master_zip_ckpt, master_addr_ckpt, geocode_ckpt, local_cache_path])
        except Exception as e:
            self.status_var.set("エラー")
            messagebox.showerror("エラー", str(e))
        finally:
            self.stop_progress_timer()
            self.run_button.config(state="normal")

    def _progress_callback(self, done: int, total: int, phase: str):
        self.progress_total = max(total, 1)
        pct = (done / self.progress_total) * 100
        self.progress_var.set(pct)
        self.status_var.set(f"ジオコーディング {done}/{total}")
        self.progress_state = {"phase": "geocode", "done": done, "total": total, "detail": "ジオコーディング"}

    def _on_master_progress(self, phase: str, done: int, total: int, detail: str):
        """マスタ突合の進捗更新"""
        self.progress_state = {"phase": phase, "done": done, "total": total, "detail": detail}
        pct = (done / max(total, 1)) * 100
        self.progress_var.set(pct)
        self.status_var.set(f"{detail} {done}/{total}")
        self.log(f"{detail} {done}/{total} ({pct:.1f}%)")

    def _on_master_addr_progress(self, col: str, done: int, total: int, detail: str):
        """住所突合の進捗更新（列単位で更新）"""
        pct = (done / max(total, 1)) * 100
        self.progress_state = {"phase": "master_addr", "done": done, "total": total, "detail": f"{detail}"}
        self.log(f"{detail} {done}/{total} ({pct:.1f}%)")
        self.progress_var.set(pct)
        self.status_var.set(f"{detail} {done}/{total}")

    def _on_output_progress(self, done: int, total: int, detail: str):
        """出力処理の進捗更新（シートコピーを含む）"""
        pct = (done / max(total, 1)) * 100
        self.progress_state = {"phase": "output", "done": done, "total": total, "detail": detail}
        self.progress_var.set(pct)
        self.status_var.set(f"{detail} {done}/{total}")

    def _build_master_highlight_indices(self) -> set:
        """マスタシートで使用した行を塗りつぶすためのindex集合を作成"""
        highlight_idx = set()
        if self.used_master_idx:
            highlight_idx.update(self.used_master_idx)
        if self.used_zip_codes:
            # 郵便番号が一致した行のindexを追加
            matched = self.master_df[self.master_df["郵便番号"].isin(self.used_zip_codes)]
            highlight_idx.update(matched.index.tolist())
        return highlight_idx

    def _cleanup_checkpoints(self, paths: list):
        """処理完了後にチェックポイントCSVを削除"""
        for p in paths:
            try:
                if p and os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass


def main():
    root = tk.Tk()
    app = GeoApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
